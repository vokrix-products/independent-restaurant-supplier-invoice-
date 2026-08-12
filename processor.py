import os
import json
import pdfplumber
import openpyxl
from typing import List, Dict, Any, Optional
from openai import OpenAI

# DeepSeek instantiation (must match brief exactly)
client = OpenAI(
    api_key=os.environ["DEEPSEEK_API_KEY"],
    base_url="https://api.deepseek.com"
)

MODEL_NAME = "deepseek-v4-flash"

STATUS_UNPROCESSED = "unprocessed:info"
MAX_RETRIES = 2

def extract_text_from_bytes(file_bytes: bytes) -> str:
    """Try PDF, then Excel, then plain text/CSV fallback."""
    # Attempt PDF via pdfplumber
    try:
        import io
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            texts = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    texts.append(text)
            if texts:
                return "\n".join(texts)
    except Exception:
        pass

    # Attempt Excel via openpyxl
    try:
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        rows = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                rows.append(",".join(row_values))
        if rows:
            return "\n".join(rows)
    except Exception:
        pass

    # Fallback: decode as UTF-8 (handles CSV/plain text)
    try:
        return file_bytes.decode("utf-8", errors="ignore")
    except Exception:
        return ""

def build_extraction_prompt(raw_text: str) -> str:
    """Construct prompt instructing DeepSeek to output invoice line items as JSON."""
    prompt = f"""
You are an invoice data extraction assistant. Extract all information from the following invoice text and return it as a single JSON object with this structure:

{{
  "supplier_name": "Supplier Company Name",
  "remit_address": "Remit-to address",
  "invoice_number": "INV-12345",
  "invoice_date": "YYYY-MM-DD",
  "due_date": "YYYY-MM-DD",
  "payment_terms": "Net 30",
  "purchase_order": "PO-98765",
  "location": "Store 1",
  "currency": "USD",
  "tax_amount": 10.00,
  "freight_shipping": 5.00,
  "invoice_total": 120.00,
  "line_items": [
    {{
      "description": "Item name",
      "sku": "SKU123",
      "unit": "kg",
      "quantity": 5,
      "unit_price": 10.00,
      "extended_total": 50.00,
      "discounts_allowances": 0.00,
      "gl_category": "Food"
    }}
  ]
}}

Important rules:
- Use ISO-8601 dates in plain YYYY-MM-DD format ONLY (e.g. 2024-07-11). NEVER include a time component or timezone offset. If a date is missing, set it to None.
- All numeric values should be floats (use None if missing).
- The primary entity being tracked is the supplier (vendor). The "title" field will later be set to supplier_name.
- If the text contains MULTIPLE invoices, return {{"invoices": [ {{...same structure...}}, {{...same structure...}} ]}} with one entry per invoice.
- If the invoice text is not an invoice or cannot be parsed, return {{}} empty object.
- If the supplier name cannot be determined, set supplier_name to null — do NOT invent a name.
- Return only the JSON object, no extra text.

Invoice text:
{raw_text}
"""
    return prompt

def parse_llm_response(response_content: str) -> dict:
    """Parse the LLM JSON response, returning a dict."""
    try:
        # Strip possible markdown fences
        if response_content.startswith("```"):
            # Remove first line and last line
            lines = response_content.split("\n")
            lines = [line for line in lines if not line.strip().startswith("```")]
            response_content = "\n".join(lines)
        data = json.loads(response_content)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def normalize_date(value) -> Optional[str]:
    """Normalize any date value to plain YYYY-MM-DD, or None."""
    if not value:
        return None
    s = str(value).strip()
    if not s or s.lower() == "none":
        return None
    # Take the date part only (handles "2024-07-11T00:00:00+00:00" etc.)
    date_part = s[:10]
    # Validate it looks like a date
    if len(date_part) == 10 and date_part[4] == "-" and date_part[7] == "-":
        return date_part
    return None

def process_file(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Extract line-item records from invoice bytes.

    Returns a list of dictionaries, each with keys:
        title (str): supplier name
        status (str): always "unprocessed:info"
        details (dict): invoice-level and line-item fields
        due_date (str or None): ISO-8601 due date
    """
    # 1. Get text from any format
    raw_text = extract_text_from_bytes(file_bytes)
    if not raw_text.strip():
        # No content -> empty list
        return []

    # 2. Prepare prompt and call DeepSeek with retries
    prompt = build_extraction_prompt(raw_text)
    for attempt in range(MAX_RETRIES + 1):
        try:
            completion = client.chat.completions.create(
                model=MODEL_NAME,
                messages=[
                    {"role": "system", "content": "You are a precise JSON invoice extractor."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            response_text = completion.choices[0].message.content
            print(f"DeepSeek attempt {attempt} raw response: {response_text[:500]}")
            invoice_data = parse_llm_response(response_text)
            if invoice_data:
                break
            print(f"DeepSeek returned empty object on attempt {attempt}")
        except Exception as e:
            print(f"DeepSeek call failed on attempt {attempt}: {e}")
            if attempt == MAX_RETRIES:
                # Return empty list on persistent failure
                return []
    else:
        # No valid response after retries
        print("DeepSeek returned empty/unparseable response on all attempts")
        return []

    # 3. Build records: one per line item, or one overall if no line items.
    #    Support both single-invoice and multi-invoice {"invoices": [...]} responses.
    records = []
    invoice_list = invoice_data.get("invoices") or [invoice_data]
    if not isinstance(invoice_list, list):
        invoice_list = [invoice_list]

    for inv in invoice_list:
        if not isinstance(inv, dict):
            continue
        supplier_name = inv.get("supplier_name")
        if not supplier_name:
            # Skip invoices with no supplier — generic placeholders violate the README
            print(f"Skipping invoice without supplier_name: {str(inv)[:200]}")
            continue
        supplier_name = str(supplier_name).strip()
        due_date = normalize_date(inv.get("due_date"))
        invoice_date = normalize_date(inv.get("invoice_date"))
        line_items = inv.get("line_items", [])

        # Common details to propagate into each record
        base_details = {
            "invoice_number": inv.get("invoice_number"),
            "invoice_date": invoice_date,
            "remit_address": inv.get("remit_address"),
            "payment_terms": inv.get("payment_terms"),
            "purchase_order": inv.get("purchase_order"),
            "location": inv.get("location"),
            "currency": inv.get("currency"),
            "tax_amount": inv.get("tax_amount"),
            "freight_shipping": inv.get("freight_shipping"),
            "invoice_total": inv.get("invoice_total"),
        }

        if line_items:
            for item in line_items:
                if not isinstance(item, dict):
                    continue
                # Merge base details with line-specific details
                details = base_details.copy()
                details.update({k: v for k, v in item.items() if v is not None})
                record = {
                    "title": supplier_name,          # primary entity
                    "status": STATUS_UNPROCESSED,    # always unprocessed on extraction
                    "details": details,
                    "due_date": due_date
                }
                records.append(record)
        else:
            # No line items: create a single record holding what we have
            details = base_details.copy()
            record = {
                "title": supplier_name,
                "status": STATUS_UNPROCESSED,
                "details": details,
                "due_date": due_date
            }
            records.append(record)

    return records
